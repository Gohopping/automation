
# coding: utf-8

# In[13]:


import requests
from hashlib import md5
from selenium import  webdriver
from time import sleep
from PIL import Image
import pyautogui  
from selenium.webdriver.support.ui import Select
import xlrd
import xlwt
from lxml import etree
import xlsxwriter
import openpyxl
from openpyxl.styles import Font,colors,Alignment
from openpyxl import Workbook,load_workbook
import xlwings as xw
import os,time

#下载路径类似这样设置  C:\Users\Administrator\Downloads
download_path = "C:\\Users\\Administrator\\Downloads"
user = "xxxxxxxx"
passwd = "xxxxxxxx"
#指定生成的表格路径类似于"E:\\"
generation_path = "E:\\"
#超级鹰自带代码部分
class Chaojiying_Client(object):

    def __init__(self, username, password, soft_id):
        self.username = username
        password =  password.encode('utf8')
        self.password = md5(password).hexdigest()
        self.soft_id = soft_id
        self.base_params = {
            'user': self.username,
            'pass2': self.password,
            'softid': self.soft_id,
        }
        self.headers = {
            'Connection': 'Keep-Alive',
            'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 5.1; Trident/4.0)',
        }

    def PostPic(self, im, codetype):
        """
        im: 图片字节
        codetype: 题目类型 参考 http://www.chaojiying.com/price.html
        """
        params = {
            'codetype': codetype,
        }
        params.update(self.base_params)
        files = {'userfile': ('ccc.jpg', im)}
        result = requests.post('http://upload.chaojiying.net/Upload/Processing.php', data=params, files=files, headers=self.headers).json()
        return result

    def ReportError(self, im_id):
        """
        im_id:报错题目的图片ID
        """
        params = {
            'id': im_id,
        }
        params.update(self.base_params)
        r = requests.post('http://upload.chaojiying.net/Upload/ReportError.php', data=params, headers=self.headers)
        return r.json()

url="http://jxgl.zjiet.edu.cn/(S(1qjwu245lbcy1qnv44jokj45))/default2.aspx"
driver=webdriver.Chrome()#打开浏览器,这里用的chrome
driver.maximize_window()#设置窗口最大化
driver.implicitly_wait(30)#设置隐式等待
driver.get(url)#打开网页
sleep(3)
driver.find_element_by_id("TextBox1").send_keys(user)#输入用户名
driver.find_element_by_id("TextBox2").send_keys(passwd)#输入用户密码

#获取验证码图片
driver.save_screenshot('a.png')
imgelement = driver.find_element_by_xpath('/html/body/form/div[3]/div[3]/dl/dd[3]/img')   #定位标签
location = imgelement.location
size = imgelement.size
rangle = (int(location['x']), int(location['y']), int(location['x'] + size['width']),
          int(location['y'] + size['height']))  # 写成我们需要截取的位置坐标
i = Image.open("a.png")  # 打开截图
frame = i.crop(rangle)  # 使用Image的crop函数，从截图中再次截取我们需要的区域
frame.save('save.png') # 保存我们接下来的验证码图片 进行打码

chaojiying = Chaojiying_Client('ahyf50', 'hyf008634', '911251')  #用户中心>>软件ID 生成一个替换 96001
im = open('save.png', 'rb').read()  #本地图片文件路径 来替换 a.jpg 有时WIN系统须要//
result = chaojiying.PostPic(im, 1004)
#print('验证码结果为:' + result['pic_str']) 

#输入验证码
driver.find_element_by_id("TextBox3").send_keys(result['pic_str'])
sleep(2)
#点击教师
driver.find_element_by_id("RadioButtonList1_1").click()
#点击登录
driver.find_element_by_id("Button1").click()

#点击回车
sleep(3)
pyautogui.hotkey("enter")

#点击信息查询
sleep(2)
driver.find_element_by_xpath("//*[@id='headDiv']/ul/li[3]/a/span").click() 

#点击选课情况
driver.find_element_by_xpath("/html/body/div/div[1]/ul/li[3]/ul/li[2]/a").click()

#定位到iframe框
driver.switch_to.frame("frame_content")

#只需要点击一下，点击选择打勾
sleep(2)
driver.find_element_by_id("CheckBox1").click()

#定义的课列表
driver.find_element_by_id("kcmc").click()
page_text = driver.page_source#获取页面
selector = etree.HTML(page_text)#解析动态页面
content = selector.xpath('//*[@id="kcmc"]/option')[:-2]#定位到班级数据位置
class_text_list = [] 
for i in content:
    class_text = i.xpath("./text()")[0]
    class_text_list.append(class_text)

#导入基于selenium的select模块
sleep(2)
for class_text in class_text_list:
    Select(driver.find_element_by_id("kcmc")).select_by_visible_text(class_text)#选择课程
    sleep(0.2)
    driver.find_element_by_id("btnPrint").click()#点击打印名单
    sleep(4)

dir_list = []
directory = download_path
t = []
d = {}
for filename in os.listdir(directory):
        path = directory + "\\" + filename
        time1 = time.ctime(os.path.getmtime(path))
        d[time1] = filename
        t.append(time1)
n = 1
for i in sorted(t):
        fn = directory + "\\" + d[i]
        new_fn = directory + "\\" + str(n)   + d[i]
        os.rename(fn, new_fn)
        n += 1
        dir_list.append(new_fn)
sleep(3)
for dir in dir_list:
    result = []#获取['任课教师：毛凌志', '课号：(2020-2021-1)-23430-2124-2', '课程名称：数据采集与清洗/考试', '开课学院：信息技术系', '学分：4.0']
    if dir.split(".")[-1] == "xls":
        workbook = xlrd.open_workbook(dir)#打开文件
        Sheet1 = workbook.sheet_by_name('Sheet1') #取出Sheet1表格
        merge_list = Sheet1.merged_cells[1:]#取出第一个合并单元格位置
        teacher_name = Sheet1.cell_value(3,3) #取出教师名字
        result.append(teacher_name)#将老师的名字放入到result列表中
        for i in merge_list:
            result.append(Sheet1.cell_value(i[0],i[2]))#循环取出合并单元格的数据并将其放入到result中

        #获取班级数据 
        Select(driver.find_element_by_id("kcmc")).select_by_visible_text(class_text_list[dir_list.index(dir)])
        sleep(2)
        page_text = driver.page_source#获取页面
        selector = etree.HTML(page_text)#解析动态页面
        content = selector.xpath('//*[@id="DataGrid1"]/tbody/tr')[1:]#定位到班级数据位置
        id_list = [] #学号列表
        name_list = [] #名字列表
        major_list = [] #专业列表
        class_name_list = [] #班级名称列表
        for i in content:
            id = i.xpath("./td[1]/text()")[0]
            name = i.xpath("./td[2]/text()")[0]
            major = i.xpath("./td[3]/text()")[0]
            class_name = i.xpath("./td[4]/text()")[0]
            id_list.append(id)
            name_list.append(name)
            major_list.append(major)
            class_name_list.append(class_name)
        sleep(2)
        course,b = result[2].split("/")
        course_name = course[5:]
        # todo 创建excel文件
        xl = xlsxwriter.Workbook(generation_path+course_name+"_"+class_name_list[-1]+"签到表.xlsx")
        Sheet = xl.add_worksheet('Sheet1')
        if len(id_list)<=50:
            Sheet.merge_range('A1:G1','签到表')
            Sheet.merge_range('A2:D2',result[1])
            Sheet.merge_range('E2:G2',result[2])
            Sheet.merge_range('A3:C3',result[3])
            Sheet.merge_range('E3:F3',result[4])
            Sheet.write_string(2,3,result[0])
            list = ['序号','学号','姓名','专业','班级','签到','备注']
            for i in range(7):
                Sheet.write_string(3,i,list[i])
                for serial_number in range(len(name_list)): 
                    if i == 0:
                        Sheet.write(serial_number+4,i,serial_number+1)
                    elif i==1:
                        Sheet.write(serial_number+4,i,id_list[serial_number])
                    elif i==2:
                        Sheet.write(serial_number+4,i,name_list[serial_number])
                    elif i==3:
                        Sheet.write(serial_number+4,i,major_list[serial_number])
                    elif i==4:
                        Sheet.write(serial_number+4,i,class_name_list[serial_number])
            xl.close()
            sleep(2)
            #签到表样式
            xl = openpyxl.load_workbook(generation_path+course_name+"_"+class_name_list[-1]+"签到表.xlsx")
            Sheet = xl['Sheet1']
            bold_16_font = Font(name="黑体",size=16,color=colors.BLACK,bold=True)
            Sheet['A1'].font = bold_16_font
            #居中样式
            center = Alignment(horizontal="center",vertical="center")
            Sheet['A1'].alignment = center
            #内容左对齐设置
            left = Alignment(horizontal="left")
            #课号样式
            class_number_9 = Font(name="Courier New",size=9)
            Sheet['A2'].font = class_number_9
            column_list = ['A','B','C','D','E','F','G']
            for row in range(len(id_list)+3):
                for column in column_list:
                    Sheet[column+str(row+2)].font = Font(name="Courier New",size=9)
                    Sheet[column+str(row+2)].alignment = left
            Sheet['F4'].font = Font(name="宋体",size=9)
            Sheet['G4'].font = Font(name="宋体",size=9)
            Sheet.column_dimensions['A'].width = 3.67+0.78
            Sheet.column_dimensions['B'].width = 10.90+0.77
            Sheet.column_dimensions['C'].width = 6.67+0.77
            Sheet.column_dimensions['D'].width = 16.67+0.77
            Sheet.column_dimensions['E'].width = 20+0.77
            Sheet.column_dimensions['F'].width = 17.11+0.77
            Sheet.column_dimensions['G'].width = 7.56+0.77
            #序号、学号、姓名、专业、班级、签到、备注居中设置
            for i in column_list:
                Sheet[i+'4'].alignment = center
            xl.save(generation_path+course_name+"_"+class_name_list[-1]+"签到表.xlsx")#最后一段要保存文件  
            sleep(2)
            #打开存好的excel
            wb = xw.Book(generation_path+course_name+"_"+class_name_list[-1]+"签到表.xlsx") #打开文件
            ws = wb.sheets['Sheet1'] #选择表格
            a_range = f'A1:{"G"}{len(id_list)+4}' #生成表格的数据范围
            #设置边框
            ws.range(a_range).api.Borders(8).LineStyle = 1 #上边框
            ws.range(a_range).api.Borders(9).LineStyle = 1 #下边框
            ws.range(a_range).api.Borders(7).LineStyle = 1 #左边框
            ws.range(a_range).api.Borders(10).LineStyle = 1 #右边框
            ws.range(a_range).api.Borders(12).LineStyle = 1 #内横边框
            ws.range(a_range).api.Borders(11).LineStyle = 1 #内纵边框
            #保存并关闭excel
            wb.save(generation_path+course_name+"_"+class_name_list[-1]+"签到表.xlsx")

        else:
            Sheet.merge_range('A1:J1','签到表')
            Sheet.merge_range('A2:E2',result[1])
            Sheet.merge_range('F2:J2',result[2])
            Sheet.merge_range('A3:E3',result[3])
            Sheet.merge_range('F3:H3',result[4])
            Sheet.merge_range('I3:J3',result[0])
            for i in range(10):
                list = ['序号','学号','姓名','班级','签到','序号','学号','姓名','班级','签到']
                Sheet.write_string(3,i,list[i])
                for serial_number in range(len(name_list)):
                    if serial_number<50 and i == 0:
                        Sheet.write(serial_number+4,i,serial_number+1)
                    elif serial_number<50 and i == 1:
                        Sheet.write(serial_number+4,i,id_list[serial_number])
                    elif serial_number<50 and i == 2:
                        Sheet.write(serial_number+4,i,name_list[serial_number])
                    elif serial_number<50 and i == 3:
                        Sheet.write(serial_number+4,i,class_name_list[serial_number])
                    elif serial_number>=50 and i == 5:
                        Sheet.write(serial_number+4-50,i,serial_number+1)
                    elif serial_number>=50 and i == 6:
                        Sheet.write(serial_number+4-50,i,id_list[serial_number])
                    elif serial_number>=50 and i == 7:
                        Sheet.write(serial_number+4-50,i,name_list[serial_number])
                    elif serial_number>=50 and i == 8:
                        Sheet.write(serial_number+4-50,i,class_name_list[serial_number])
            #todo 关闭文件
            xl.close()
            sleep(1)
            #创建excel表格
            xl = openpyxl.load_workbook(generation_path+course_name+"_"+class_name_list[-1]+"签到表.xlsx")
            Sheet = xl['Sheet1']
            #签到表样式
            bold_16_font = Font(name="黑体",size=16,color=colors.BLACK,bold=True)
            Sheet['A1'].font = bold_16_font
            #居中样式
            center = Alignment(horizontal="center",vertical="center")
            Sheet['A1'].alignment = center
            #内容左对齐设置
            left = Alignment(horizontal="left")
            #课号样式
            class_number_9 = Font(name="Courier New",size=9)
            Sheet['A2'].font = class_number_9
            column_list = ['A','B','C','D','E','F','G','H','I','J']
            for row in range(len(id_list)+3):
                for column in column_list:
                    Sheet[column+str(row+2)].font = Font(name="Courier New",size=9)
                    Sheet[column+str(row+2)].alignment = left
            Sheet['E4'].font = Font(name="宋体",size=9)
            Sheet['J4'].font = Font(name="宋体",size=9)
            Sheet['G2'].font = Font(name="宋体",size=9)
            Sheet['A3'].font = Font(name="宋体",size=9)
            Sheet.column_dimensions['A'].width = 3.67+0.78
            Sheet.column_dimensions['B'].width = 10.90+0.77
            Sheet.column_dimensions['C'].width = 5.22+0.77
            Sheet.column_dimensions['D'].width = 9.33+0.77
            Sheet.column_dimensions['E'].width = 12.44+0.77
            Sheet.column_dimensions['F'].width = 3.67+0.78
            Sheet.column_dimensions['G'].width = 10.90+0.77
            Sheet.column_dimensions['H'].width = 5.22+0.77
            Sheet.column_dimensions['I'].width = 9.33+0.77
            Sheet.column_dimensions['J'].width = 12.44+0.77
            #序号、学号、姓名、专业、班级、签到、备注居中设置
            for i in column_list:
                Sheet[i+'4'].alignment = center
            Sheet['A2'].alignment = center
            Sheet['F2'].alignment = center
            Sheet['A3'].alignment = center
            Sheet['F3'].alignment = center
            Sheet['I3'].alignment = center
            xl.save(generation_path+course_name+"_"+class_name_list[-1]+"签到表.xlsx")#最后一段要保存文件
            sleep(1)
            #打开存好的excel
            wb = xw.Book(generation_path+course_name+"_"+class_name_list[-1]+"签到表.xlsx") #打开文件
            ws = wb.sheets['Sheet1'] #选择表格
            a_range = f'A1:{"J"}{54}' #生成表格的数据范围
            #设置边框
            ws.range(a_range).api.Borders(8).LineStyle = 1 #上边框
            ws.range(a_range).api.Borders(9).LineStyle = 1 #下边框
            ws.range(a_range).api.Borders(7).LineStyle = 1 #左边框
            ws.range(a_range).api.Borders(10).LineStyle = 1 #右边框
            ws.range(a_range).api.Borders(12).LineStyle = 1 #内横边框
            ws.range(a_range).api.Borders(11).LineStyle = 1 #内纵边框
            #保存并关闭excel
            wb.save(generation_path+course_name+"_"+class_name_list[-1]+"签到表.xlsx")

